import json
import sys
from pathlib import Path


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def load_workbook(workbook_path):
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl이 필요합니다. 터미널에서 `py -m pip install openpyxl` 실행 후 다시 시도하세요."
        ) from exc

    return openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)


def get_sheet(wb, names):
    for name in names:
        if name in wb.sheetnames:
            return wb[name]

    raise RuntimeError(
        "엑셀에 필요한 시트가 없습니다.\n"
        f"찾은 시트: {', '.join(wb.sheetnames)}\n"
        f"필요한 후보: {', '.join(names)}"
    )


def convert_basic_data_sheet(ws, pc_name, id_index=1, sol_index=2, fumen_index=3, imgur_index=4):
    rows = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        setup_id = clean_text(row[id_index] if len(row) > id_index else "")
        sol = row[sol_index] if len(row) > sol_index else None
        fumen = clean_text(row[fumen_index] if len(row) > fumen_index else "")
        imgur = clean_text(row[imgur_index] if len(row) > imgur_index else "")

        if not setup_id:
            continue

        rows.append(
            {
                "id": setup_id.upper(),
                "sol": sol,
                "fumen": fumen,
                "imgur": imgur,
                "pc": pc_name,
            }
        )

    return rows


def convert_third_data_sheet(ws):
    return convert_basic_data_sheet(
        ws,
        "third",
        id_index=1,
        sol_index=2,
        fumen_index=4,
        imgur_index=6,
    )


def convert_fourth_data_sheet(ws):
    rows = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        setup_id = clean_text(row[2] if len(row) > 2 else "")
        pieces = clean_text(row[3] if len(row) > 3 else "").upper()
        held = clean_text(row[4] if len(row) > 4 else "").upper()
        sol = row[5] if len(row) > 5 else None
        fumen = clean_text(row[7] if len(row) > 7 else "")
        imgur = clean_text(row[13] if len(row) > 13 else "")

        if not setup_id or not pieces or not held:
            continue

        rows.append(
            {
                "id": setup_id.upper(),
                "pieces": pieces,
                "held": held,
                "sol": sol,
                "fumen": fumen,
                "imgur": imgur,
                "pc": "fourth",
            }
        )

    return rows


def convert_fifth_data_sheet(ws):
    rows = []

    for row in ws.iter_rows(min_row=36, values_only=True):
        base = clean_text(row[1] if len(row) > 1 else "").upper()
        piece_count = row[3] if len(row) > 3 else None
        setup_id = clean_text(row[4] if len(row) > 4 else "").upper()
        imgur = clean_text(row[5] if len(row) > 5 else "")
        fumen = clean_text(row[7] if len(row) > 7 else "")
        sol = row[9] if len(row) > 9 else None

        if not base or not setup_id or piece_count in (None, ""):
            continue

        rows.append(
            {
                "id": setup_id,
                "base": base,
                "piece_count": int(float(piece_count)),
                "sol": sol,
                "fumen": fumen,
                "imgur": imgur,
                "pc": "fifth",
            }
        )

    return rows


def convert_sixth_data_sheet(ws):
    rows = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        setup_id = clean_text(row[1] if len(row) > 1 else "").upper()
        setup_id2 = clean_text(row[2] if len(row) > 2 else "").upper()
        sol = row[3] if len(row) > 3 else None
        fumen = clean_text(row[11] if len(row) > 11 else "")
        imgur = clean_text(row[13] if len(row) > 13 else "")

        if not setup_id or not setup_id2:
            continue

        rows.append(
            {
                "id": setup_id,
                "id2": setup_id2,
                "sol": sol,
                "fumen": fumen,
                "imgur": imgur,
                "pc": "sixth",
            }
        )

    return rows


def main():
    base_dir = Path(__file__).resolve().parents[2]

    if len(sys.argv) >= 2:
        workbook_path = Path(sys.argv[1]).resolve()
    else:
        workbook_path = base_dir / "Setup Konbini.xlsx"

    if not workbook_path.exists():
        raise RuntimeError(f"엑셀 파일을 찾지 못했습니다: {workbook_path}")

    wb = load_workbook(workbook_path)

    first_ws = get_sheet(wb, ["1st: Data", "1st Data", "1st:Data"])
    second_ws = get_sheet(wb, ["2nd: Data", "2nd Data", "2nd:Data"])
    third_ws = get_sheet(wb, ["3rd: Data", "3rd Data", "3rd:Data"])
    fourth_ws = get_sheet(wb, ["4th: Data", "4th Data", "4th:Data"])
    fifth_ws = get_sheet(wb, ["5th: Data", "5th Data", "5th:Data"])
    sixth_ws = get_sheet(wb, ["6th: Data", "6th Data", "6th:Data"])
    seventh_ws = get_sheet(wb, ["7th: Data", "7th Data", "7th:Data"])

    data = {
        "version": 3,
        "source": workbook_path.name,
        "first": convert_basic_data_sheet(first_ws, "first"),
        "second": convert_basic_data_sheet(
            second_ws,
            "second",
            id_index=1,
            sol_index=2,
            fumen_index=4,
            imgur_index=6,
        ),
        "third": convert_third_data_sheet(third_ws),
        "fourth": convert_fourth_data_sheet(fourth_ws),
        "fifth": convert_fifth_data_sheet(fifth_ws),
        "sixth": convert_sixth_data_sheet(sixth_ws),
        "seventh": convert_basic_data_sheet(seventh_ws, "seventh"),
    }

    out_path = Path(__file__).resolve().parent / "setup_data.json"
    out_path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"완료: {out_path}")
    print(f"1st Data rows: {len(data['first'])}")
    print(f"2nd Data rows: {len(data['second'])}")
    print(f"3rd Data rows: {len(data['third'])}")
    print(f"4th Data rows: {len(data['fourth'])}")
    print(f"5th Data rows: {len(data['fifth'])}")
    print(f"6th Data rows: {len(data['sixth'])}")
    print(f"7th Data rows: {len(data['seventh'])}")


if __name__ == "__main__":
    main()
